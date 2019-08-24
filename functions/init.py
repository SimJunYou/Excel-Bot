from openpyxl import load_workbook
import logging, redis, os

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)


logger.info("Opening Excel file")
main_workbook = load_workbook('SeminarDatasheet.xlsx')
ws = main_workbook['Sheet1']

rList = redis.from_url(os.environ.get("REDIS_URL"))
rList.delete('Feedback')  # clearing feedback from tests

PERSON = []  # PERSON for local work, redis for heroku database
logger.info("Excel file dumped into working list and redis")

main_workbook.save('SeminarDatasheet.xlsx')
logger.info("Excel file closed")

row_number = 2
while ws['A' + str(row_number)].value is not None:
    NRIC = ws['A' + str(row_number)].value
    GRP_ID = ws['B' + str(row_number)].value
    PERSON.append({'NRIC': NRIC,
                   'GRP1': GRP_ID,
                   'GRP1_REG': ''})
    rList.mset({NRIC: ''})  # dump NRIC into redis
    row_number += 1


TYPING_NRIC, ENDSEM = range(2)  # for conv_handler
QN2, QN3, ENDPOST = range(3)  # for post_conv_handler
ADMIN_START, ADMIN_END = range(2)  # for admin_handler

adminID = [234058962, 460007829]

# # initialise ADMIN_LIST if not already done
# if not rList.get("ADMIN_LIST"):
#     rList.mset({"ADMIN_LIST": [234058962, 460007829]})


# MAIN VARIABLES ARE: ws, rList, PERSON, TYPING_NRIC & RESPONSE, QN1 & QN2 & QN3, adminID
