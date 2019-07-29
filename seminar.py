#!/usr/bin/env python
# -*- coding: utf-8 -*-

from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (Updater, CommandHandler, MessageHandler, Filters, RegexHandler,
                          ConversationHandler)

from openpyxl import load_workbook
from excel import returnSeating, createFile

import logging
import os, redis
import string, random

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)

# welcome to my Macbook!

# <EXCEL AND REDIS>
logger.info("Opening Excel file")
main_workbook = load_workbook('SeminarDatasheet.xlsx')
ws = main_workbook['Sheet1']

rList = redis.from_url(os.environ.get("REDIS_URL"))
PERSON = []  # PERSON for local work, redis for heroku database

row_number = 2
while ws['A' + str(row_number)].value is not None:
    NRIC = ws['A' + str(row_number)].value
    GRP_ID = ws['B' + str(row_number)].value
    PERSON.append({'NRIC': NRIC,
                   'GRP1': GRP_ID,
                   'GRP1_REG': ''})
    rList.mset({NRIC: ''})  # dump NRIC into redis
    row_number += 1

logger.info("Excel file dumped into working list and redis")
main_workbook.save('SeminarDatasheet.xlsx')
logger.info("Excel file closed")

# </EXCEL AND REDIS>


# init values for Telegram
TYPING_NRIC, RESPONSE = range(2)
reply_keyboard = [['Yes', 'No']]
markup = ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True)
remove = ReplyKeyboardRemove(remove_keyboard=True)

# init the kill switch! ^$ is regex for start and end with, meaning perfect match
killSwitch = '^' + ''.join(random.choices(string.ascii_uppercase, k=15)) + '$'

def validate_nric(nric):
    if len(nric) == 5:
        if nric[:4].isdigit() and nric[4].isalpha():
            return True
    return False


###############


def start(bot, update):
    update.message.reply_text(
        ''' _Welcome to the Redesign Seminar!_
*============================*
I'm here to mark your attendance and provide your assigned group. You may leave this chat at any time, it will not affect the bot.

Please enter the _last 5 characters_ of your NRIC:''',
        parse_mode='Markdown')

    logger.info("User %s initiates contact", update.message.from_user.first_name)

    return TYPING_NRIC


def get_nric(bot, update, user_data):
    text = update.message.text
    user_data['NRIC'] = text.upper()
    if validate_nric(text):
        update.message.reply_text(
            "To confirm, the last 5 characters of your NRIC are {}.\n\nIs this correct? Yes/No".format(text),
            reply_markup=markup)

        return RESPONSE

    else:
        update.message.reply_text(
            "*The provided partial NRIC {} is incorrect!* Please check that it is in the format 1234E (where full NRIC would be S9951234E)\n\nLet's try again. Last 5 characters of your NRIC:".format(
                text),
            parse_mode='Markdown')

        return TYPING_NRIC


def final(bot, update, user_data):
    text = update.message.text
    if text.lower() == "yes":
        # MAIN ACTION
        seating = returnSeating(PERSON, user_data['NRIC'], rList)
        if seating is None:
            update.message.reply_text(
                "We cannot find your NRIC, please look for assistance around the venue.\n\nYou may now leave this chat or type /start to register another NRIC entry",
                reply_markup=remove)
        else:
            seating = seating.upper()
            update.message.reply_text(
                '''Your assigned group is: {}.\n\nThank you for attending this seminar! You may now leave this chat or type /start to register another NRIC entry.'''.format(
                    seating),
                reply_markup=remove)
    elif text.lower() == "no":
        update.message.reply_text("Let's try again. Enter the last 5 digits of your NRIC:", reply_markup = remove)
        return TYPING_NRIC

    logger.info("User {} completes".format(update.message.from_user.first_name))
    user_data.clear()
    return ConversationHandler.END


###############

# admin stuff
def collectStats(bot, update):
    if update.message.chat_id == 234058962:
        count = 0
        total = 0
        for each in PERSON:
            total += 1
            if each['GRP1_REG'] == 'P':
                count += 1
        update.message.reply_text("Good day, admin.\nTotal: {}, Present: {}".format(total, count))
        logger.info("Admin initiates stats report.\nTotal: {}, Present: {}".format(total, count))
    else:
        update.message.reply_text("You are not recognised!")


def sendFile(bot, update):
    if update.message.chat_id == 234058962:
        update.message.reply_text("Good day, admin")
        logger.info("Admin requests latest file")
        createFile(ws, rList)
        bot.send_document(update.message.chat_id, document=open('Attendance.xlsx', 'rb'))
    else:
        update.message.reply_text("You are not recognised!")


def killCommand(bot, update):
    if update.message.chat_id == 234058962:
        update.message.reply_text("Are you absolutely certain you want to kill the bot?")
        update.message.reply_text("Type the kill-code to kill the bot: "+killSwitch[1:-1])
        logger.info("Admin initiates kill command, requesting confirmation")
    else:
        update.message.reply_text("You are not recognised!")


def error(bot, update, error):
    """Log Errors caused by Updates."""
    logger.warning('Update "%s" caused error "%s"', update, error)


###############

def main():
    # Create the Updater and pass it your bot's token.
    updater = Updater("241346491:AAH09_cf9KfaFohGgXUo96ljvOeyqcD1k4o")

    # Get the dispatcher to register handlers
    dp = updater.dispatcher

    # Add conversation handler with the states GENDER, PHOTO, LOCATION and BIO
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            TYPING_NRIC: [MessageHandler(Filters.text,
                                         get_nric,
                                         pass_user_data=True)],
            RESPONSE: [RegexHandler('^Yes|No$',
                                    final,
                                    pass_user_data=True)]
        },
        fallbacks=[]
    )


    def killBot(bot, update):
        chat_id = update.message.chat_id
        if update.message.chat_id == 234058962:
            update.message.reply_text("Running createFile subroutine...")
            createFile(rList)
            bot.send_document(chat_id, document=open('Attendance.xlsx', 'rb'))
            update.message.reply_text("Bot is being killed. Goodbye, admin.")
            logger.info("Bot has been killed.")
            updater.stop()
            updater.is_idle = False
        else:
            update.message.reply_text("You are not recognised!")

    dp.add_handler(conv_handler)
    dp.add_handler(CommandHandler('kill', killCommand))
    dp.add_handler(CommandHandler('stats', collectStats))
    dp.add_handler(CommandHandler('file', sendFile))
    dp.add_handler(RegexHandler(killSwitch, killBot))

    # log all errors
    dp.add_error_handler(error)

    # Start the Bot
    updater.start_polling()
    logger.info("The bot is running.")
    # Run the bot until you press Ctrl-C or the process receives SIGINT,
    # SIGTERM or SIGABRT. This should be used most of the time, since
    # start_polling() is non-blocking and will stop the bot gracefully.
    updater.idle()


if __name__ == '__main__':
    main()
